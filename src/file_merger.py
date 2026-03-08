"""
File Merger - Core Module

This module contains the core logic for merging Excel and CSV files from subdirectories.
"""

import os
import re
import sys
import glob
import pandas as pd
from pathlib import Path
from typing import List, Dict, Tuple, Optional, Any

try:
    from .structure_validator import (
        validate_column_structure,
        format_validation_report,
        StructureValidationResult,
        file_has_plural_singular_conflict,
        compute_canonical_plural_singular_renames,
    )
except ImportError:
    from structure_validator import (
        validate_column_structure,
        format_validation_report,
        StructureValidationResult,
        file_has_plural_singular_conflict,
        compute_canonical_plural_singular_renames,
    )


class FileMerger:
    """Handles merging of Excel and CSV files within subdirectories."""
    
    # Default threshold (MB) above which chunked/streaming read is used. None = always use standard read.
    DEFAULT_LARGE_FILE_THRESHOLD_MB = 50.0

    def __init__(self, root_directory: str, clean_columns: bool = True,
                 column_cleaning_options: Dict[str, bool] = None,
                 validate_structure: bool = True,
                 columns_to_export: Optional[List[str]] = None,
                 large_file_threshold_mb: Optional[float] = None,
                 subdir_include: Optional[List[str]] = None,
                 subdir_exclude: Optional[List[str]] = None,
                 group_by: Optional[List[str]] = None,
                 max_rows_before_aggregate: Optional[int] = None,
                 aggregate_columns: Optional[Dict[str, str]] = None,
                 aggregate_default: Optional[str] = None,
                 subdirectory_config: Optional[Dict[str, Dict[str, Any]]] = None,
                 sort_by: Optional[List[str]] = None,
                 sort_ascending: bool = False,
                 column_aliases: Optional[Dict[str, str]] = None,
                 column_canonical_patterns: Optional[List[Dict[str, str]]] = None):
        """
        Initialize the Excel merger.

        Args:
            root_directory: Path to the root directory containing subdirectories
            clean_columns: Whether to clean column names
            column_cleaning_options: Dictionary of cleaning options
            validate_structure: If True, validate column structure before merge and fail on mismatch
            columns_to_export: If set, only these columns are read and exported (from YAML or CLI)
            large_file_threshold_mb: If set, files at or above this size (MB) use chunked/streaming read.
                None disables adaptive reading (default). Use DEFAULT_LARGE_FILE_THRESHOLD_MB for 50 MB.
            subdir_include: If set, only process subdirectories whose names are in this list.
            subdir_exclude: If set, skip subdirectories whose names are in this list.
            group_by: Column names to group by when aggregating (used with max_rows_before_aggregate).
            max_rows_before_aggregate: If merged row count exceeds this and group_by is set, aggregate by group_by.
            aggregate_columns: Optional dict column_name -> agg_func (e.g. {"Impressions": "sum", "CTR": "mean"}).
                Supported: sum, mean, count, first, last, min, max, median, std.
            aggregate_default: Internal; columns not in aggregate_columns use sum (numeric) or first (other).
            subdirectory_config: Optional per-subdir overrides.
            sort_by: Optional list of column names to sort the merged output by (applied before save).
            sort_ascending: If True, sort ascending; if False (default), sort descending (e.g. largest first).
            column_aliases: Optional dict variant_name -> canonical_name for normalizing column names.
            column_canonical_patterns: Optional list of {base, canonical} to map e.g. "Name (suffix)" -> canonical.
        """
        self.root_directory = Path(root_directory)
        self.validate_structure = validate_structure
        self.columns_to_export = columns_to_export
        self.large_file_threshold_mb = large_file_threshold_mb
        self.subdir_include = subdir_include
        self.subdir_exclude = subdir_exclude or []
        self.group_by = group_by
        self.max_rows_before_aggregate = max_rows_before_aggregate
        self.aggregate_columns = aggregate_columns or {}
        self.aggregate_default = aggregate_default or None
        self.subdirectory_config = subdirectory_config or {}
        self.sort_by = sort_by or []
        self.sort_ascending = sort_ascending
        self.column_aliases = column_aliases or {}
        self.column_canonical_patterns = column_canonical_patterns or []
        self.processed_subdirs = 0
        self.total_files_merged = 0
        self.skipped_single_file = 0
        self.errors = []
        self.clean_columns = clean_columns
        
        # Default column cleaning options
        self.column_cleaning_options = {
            'strip_whitespace': True,
            'normalize_spaces': True,
            'lowercase': False,
            'remove_special_chars': False,
            'handle_common_variations': True,
            'normalize_plural_singular': False  # Off by default; only validation reports plural/singular hints
        }
        
        # Update with user-provided options
        if column_cleaning_options:
            self.column_cleaning_options.update(column_cleaning_options)
        
    def validate_directory(self) -> bool:
        """
        Validate that the root directory exists and contains subdirectories.
        
        Returns:
            bool: True if valid, False otherwise
        """
        if not self.root_directory.exists():
            self.errors.append(f"Directory does not exist: {self.root_directory}")
            return False
            
        if not self.root_directory.is_dir():
            self.errors.append(f"Path is not a directory: {self.root_directory}")
            return False
            
        # Check if there are any subdirectories
        subdirs = [d for d in self.root_directory.iterdir() if d.is_dir()]
        if not subdirs:
            self.errors.append(f"No subdirectories found in: {self.root_directory}")
            return False
            
        return True
    
    def get_supported_files(self, directory: Path) -> List[Path]:
        """
        Get all supported files (Excel and CSV) in a directory.
        Excludes Excel lock files (names starting with ~$).

        Args:
            directory: Directory to search for files

        Returns:
            List of supported file paths
        """
        supported_extensions = ['*.xlsx', '*.xls', '*.csv']
        supported_files = []

        for extension in supported_extensions:
            supported_files.extend(directory.glob(extension))

        return [p for p in supported_files if not p.name.startswith('~$')]
    
    def _detect_file_format(self, file_path: Path) -> str:
        """
        Detect file format based on extension.

        Args:
            file_path: Path to the file

        Returns:
            'excel' for .xlsx/.xls files, 'csv' for .csv files
        """
        file_ext = file_path.suffix.lower()

        if file_ext in ['.xlsx', '.xls']:
            return 'excel'
        elif file_ext == '.csv':
            return 'csv'
        else:
            return 'unknown'

    def _is_large_file(self, file_path: Path) -> bool:
        """
        Return True if the file size is at or above the configured large-file threshold.

        Args:
            file_path: Path to the file

        Returns:
            True if adaptive (chunked/streaming) read should be used
        """
        if self.large_file_threshold_mb is None:
            return False
        try:
            size_mb = file_path.stat().st_size / (1024 * 1024)
            return size_mb >= self.large_file_threshold_mb
        except OSError:
            return False
    
    def _to_singular(self, name: str) -> str:
        """
        Simple heuristic: plural form to singular (for column name normalization).
        Only used when normalize_plural_singular is enabled; does not handle all English plurals.
        """
        if not name or len(name) < 2:
            return name
        if name.endswith('ies') and len(name) > 3:
            return name[:-3] + 'y'
        if name.endswith('es') and len(name) > 2:
            return name[:-2]
        if name.endswith('s') and len(name) > 1:
            return name[:-1]
        return name

    def clean_column_names(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Clean column names in a DataFrame.
        
        Args:
            df: DataFrame to clean
            
        Returns:
            DataFrame with cleaned column names
        """
        if not self.clean_columns or df.empty:
            return df

        original_columns = df.columns.tolist()
        original_set = set(original_columns)
        cleaned_columns = []
        
        for col in original_columns:
            cleaned_col = str(col)
            
            # Strip leading and trailing whitespace
            if self.column_cleaning_options.get('strip_whitespace', True):
                cleaned_col = cleaned_col.strip()
            
            # Normalize multiple spaces to single space
            if self.column_cleaning_options.get('normalize_spaces', True):
                cleaned_col = re.sub(r'\s+', ' ', cleaned_col)
            
            # Convert to lowercase
            if self.column_cleaning_options.get('lowercase', False):
                cleaned_col = cleaned_col.lower()
            
            # Remove special characters
            if self.column_cleaning_options.get('remove_special_chars', False):
                cleaned_col = re.sub(r'[^\w\s-]', '', cleaned_col)
            
            # Handle common variations
            if self.column_cleaning_options.get('handle_common_variations', True):
                # Common replacements
                variations = {
                    'id': 'ID',
                    'date': 'Date',
                    'name': 'Name',
                    'email': 'Email',
                    'phone': 'Phone',
                    'address': 'Address',
                    'city': 'City',
                    'state': 'State',
                    'zip': 'ZIP',
                    'zipcode': 'ZIP',
                    'zip_code': 'ZIP'
                }
                
                for old, new in variations.items():
                    if cleaned_col.lower() == old.lower():
                        cleaned_col = new
                        break
            
            # Normalize plural to singular only when singular is not already a column (avoid collapsing)
            if self.column_cleaning_options.get('normalize_plural_singular', True):
                singular = self._to_singular(cleaned_col)
                if singular != cleaned_col and singular not in original_set:
                    cleaned_col = singular

            cleaned_columns.append(cleaned_col)

        # Apply canonical mapping: aliases (exact) then patterns (e.g. "Name (suffix)" -> canonical)
        for i, col in enumerate(cleaned_columns):
            if col in self.column_aliases:
                cleaned_columns[i] = self.column_aliases[col]
            else:
                for rule in self.column_canonical_patterns:
                    base = rule.get('base') or ''
                    canonical = rule.get('canonical')
                    if not base or canonical is None:
                        continue
                    pattern = re.compile(r'^' + re.escape(base) + r'\s*\([^)]*\)\s*$')
                    if pattern.match(col):
                        cleaned_columns[i] = canonical
                        break

        # Create new DataFrame with cleaned column names
        df_cleaned = df.copy()
        df_cleaned.columns = cleaned_columns
        
        # Log column changes if any
        changes = []
        for orig, cleaned in zip(original_columns, cleaned_columns):
            if orig != cleaned:
                changes.append(f"'{orig}' → '{cleaned}'")
        
        if changes:
            print(f"    📝 Column names cleaned: {', '.join(changes)}")
        
        return df_cleaned
    
    def read_file(self, file_path: Path) -> Optional[pd.DataFrame]:
        """
        Read a supported file (Excel or CSV) and return a DataFrame.

        Uses adaptive reading: when file size is at or above large_file_threshold_mb,
        switches to chunked (CSV) or streaming (Excel .xlsx) for memory efficiency.
        """
        file_format = self._detect_file_format(file_path)
        is_large = self._is_large_file(file_path)

        if file_format == 'excel':
            if is_large and file_path.suffix.lower() == '.xlsx':
                return self._read_excel_streaming(file_path)
            return self.read_excel_file(file_path)
        elif file_format == 'csv':
            if is_large:
                return self._read_with_csv_chunked(file_path)
            return self._read_with_csv(file_path)
        else:
            self.errors.append(f"Unsupported file format: {file_path}")
            return None
    
    def read_excel_file(self, file_path: Path) -> Optional[pd.DataFrame]:
        """
        Read an Excel file and return a DataFrame.
        
        Args:
            file_path: Path to the Excel file
            
        Returns:
            DataFrame or None if reading fails
        """
        # Try multiple methods in order of preference
        methods = [
            ("pandas with engine detection", self._read_with_pandas),
            ("xlwings", self._read_with_xlwings)
        ]
        
        for method_name, method_func in methods:
            try:
                df = method_func(file_path)
                if df is not None:
                    return df
            except Exception as e:
                # Continue to next method
                continue
        
        # If all methods failed, add error
        self.errors.append(f"Error reading {file_path}: All reading methods failed")
        return None
    
    def read_file_with_method(self, file_path: Path) -> Tuple[Optional[pd.DataFrame], str]:
        """
        Read a supported file and return a DataFrame with the method used.

        Uses adaptive reading for large files (chunked CSV or streaming Excel).
        """
        file_format = self._detect_file_format(file_path)
        is_large = self._is_large_file(file_path)

        if file_format == 'excel':
            if is_large and file_path.suffix.lower() == '.xlsx':
                try:
                    df = self._read_excel_streaming(file_path)
                    return (df, "openpyxl read_only (streaming)") if df is not None else (None, "failed")
                except Exception as e:
                    self.errors.append(f"Error reading Excel (streaming) {file_path}: {str(e)}")
                    return None, "failed"
            return self.read_excel_file_with_method(file_path)
        elif file_format == 'csv':
            if is_large:
                try:
                    df = self._read_with_csv_chunked(file_path)
                    return (df, "pandas CSV (chunked)") if df is not None else (None, "failed")
                except Exception as e:
                    self.errors.append(f"Error reading CSV (chunked) {file_path}: {str(e)}")
                    return None, "failed"
            try:
                df = self._read_with_csv(file_path)
                if df is not None:
                    return df, "pandas CSV reader"
                return None, "failed"
            except Exception as e:
                self.errors.append(f"Error reading CSV {file_path}: {str(e)}")
                return None, "failed"
        else:
            self.errors.append(f"Unsupported file format: {file_path}")
            return None, "failed"
    
    def read_excel_file_with_method(self, file_path: Path) -> Tuple[Optional[pd.DataFrame], str]:
        """
        Read an Excel file and return a DataFrame with the method used.
        
        Args:
            file_path: Path to the Excel file
            
        Returns:
            Tuple of (DataFrame or None, method_name)
        """
        # Try multiple methods in order of preference
        methods = [
            ("pandas with engine detection", self._read_with_pandas),
            ("xlwings", self._read_with_xlwings)
        ]
        
        for method_name, method_func in methods:
            try:
                df = method_func(file_path)
                if df is not None:
                    return df, method_name
            except Exception as e:
                # Continue to next method
                continue
        
        # If all methods failed, add error
        self.errors.append(f"Error reading {file_path}: All reading methods failed")
        return None, "failed"
    
    # Row block size for memory-efficient Excel reads (openpyxl read_only)
    EXCEL_ROW_BLOCK_SIZE = 50_000

    def _read_excel_streaming(self, file_path: Path) -> Optional[pd.DataFrame]:
        """
        Read a large .xlsx file using openpyxl read_only mode to limit memory use.

        Only supports .xlsx. Returns a single DataFrame with cleaned column names.
        """
        from openpyxl import load_workbook

        wb = None
        try:
            wb = load_workbook(str(file_path), read_only=True, data_only=True)
            ws = wb.active
            if ws is None:
                return None
            rows_iter = ws.iter_rows(values_only=True)
            headers = next(rows_iter)
            if not headers:
                return None
            headers = [str(h) if h is not None else "" for h in headers]
            ncols = len(headers)
            chunks = []
            block = []
            for row in rows_iter:
                row = list(row)[:ncols]
                if len(row) < ncols:
                    row.extend([None] * (ncols - len(row)))
                block.append(row)
                if len(block) >= self.EXCEL_ROW_BLOCK_SIZE:
                    chunks.append(pd.DataFrame(block, columns=headers))
                    block = []
            if block:
                chunks.append(pd.DataFrame(block, columns=headers))
        except Exception:
            return None
        finally:
            if wb is not None:
                wb.close()

        if not chunks:
            return self.clean_column_names(pd.DataFrame(columns=headers)) if headers else None
        df = pd.concat(chunks, ignore_index=True)
        return self.clean_column_names(df)

    def _read_with_pandas(self, file_path: Path) -> Optional[pd.DataFrame]:
        """Read Excel file using pandas with appropriate engine."""
        file_ext = file_path.suffix.lower()

        if file_ext == '.xlsx':
            # Use openpyxl for .xlsx files
            df = pd.read_excel(file_path, engine='openpyxl')
        elif file_ext == '.xls':
            # Try xlrd first for .xls files, fallback to openpyxl if it fails
            try:
                df = pd.read_excel(file_path, engine='xlrd')
            except Exception:
                # Some .xls files are actually XML-based Excel files
                df = pd.read_excel(file_path, engine='openpyxl')
        else:
            # Try to auto-detect
            df = pd.read_excel(file_path)

        # Clean column names
        return self.clean_column_names(df)
    
    def _read_with_xlwings(self, file_path: Path) -> Optional[pd.DataFrame]:
        """Read Excel file using xlwings as fallback (macOS only)."""
        try:
            import xlwings as xw
            
            # Open the workbook
            wb = xw.Book(str(file_path))
            
            # Get the first worksheet
            ws = wb.sheets[0]
            
            # Read all data from the worksheet
            data = ws.used_range.value
            
            # Close the workbook
            wb.close()
            
            if data is None:
                return None
            
            # Convert to DataFrame
            if isinstance(data, list) and len(data) > 0:
                # First row as headers
                headers = data[0]
                rows = data[1:]
                
                # Create DataFrame
                df = pd.DataFrame(rows, columns=headers)
                # Clean column names
                return self.clean_column_names(df)
            else:
                return None
                
        except ImportError:
            # xlwings not available (e.g., in Linux containers)
            raise ImportError("xlwings not available in this environment")
        except Exception as e:
            # xlwings failed, return None to try next method
            raise e
    
    # Chunk size for memory-efficient CSV reads (number of rows per chunk)
    CSV_CHUNK_SIZE = 50_000

    def _detect_csv_options(self, file_path: Path, nrows: int = 500) -> Optional[Tuple[str, str]]:
        """
        Detect encoding and delimiter by probing the first nrows of a CSV file.

        Returns:
            (encoding, delimiter) or None if no combination succeeded
        """
        encodings = ['utf-8', 'latin-1', 'windows-1252', 'utf-16']
        delimiters = [',', ';', '\t', '|']
        for encoding in encodings:
            for delimiter in delimiters:
                try:
                    df = pd.read_csv(
                        file_path,
                        encoding=encoding,
                        delimiter=delimiter,
                        quotechar='"',
                        skipinitialspace=True,
                        on_bad_lines='skip',
                        nrows=nrows,
                    )
                    if len(df.columns) > 1 and len(df) > 0:
                        return (encoding, delimiter)
                except Exception:
                    continue
        return None

    def _read_with_csv(self, file_path: Path) -> Optional[pd.DataFrame]:
        """
        Read CSV file using pandas with automatic detection.

        Args:
            file_path: Path to the CSV file

        Returns:
            DataFrame or None if reading fails
        """
        # Common encodings to try
        encodings = ['utf-8', 'latin-1', 'windows-1252', 'utf-16']

        # Common delimiters to try
        delimiters = [',', ';', '\t', '|']

        # Try different combinations of encoding and delimiter
        for encoding in encodings:
            for delimiter in delimiters:
                try:
                    df = pd.read_csv(
                        file_path,
                        encoding=encoding,
                        delimiter=delimiter,
                        quotechar='"',
                        skipinitialspace=True,
                        on_bad_lines='skip'  # Skip problematic lines
                    )

                    # Check if we got a reasonable result
                    if len(df.columns) > 1 and len(df) > 0:
                        return self.clean_column_names(df)

                except Exception:
                    continue

        # If all combinations failed, try pandas auto-detection
        try:
            df = pd.read_csv(file_path, encoding='utf-8', on_bad_lines='skip')
            return self.clean_column_names(df)
        except Exception:
            # Last resort: try with different quote characters
            try:
                df = pd.read_csv(
                    file_path,
                    encoding='utf-8',
                    quotechar="'",
                    on_bad_lines='skip'
                )
                return self.clean_column_names(df)
            except Exception:
                return None

    def _read_with_csv_chunked(self, file_path: Path) -> Optional[pd.DataFrame]:
        """
        Read a large CSV file in chunks to limit memory use.

        Uses _detect_csv_options for encoding/delimiter, then reads in chunks
        and concatenates. Returns a single DataFrame with cleaned column names.
        """
        opts = self._detect_csv_options(file_path)
        if opts is None:
            opts = ('utf-8', ',')
        encoding, delimiter = opts

        chunks = []
        try:
            reader = pd.read_csv(
                file_path,
                encoding=encoding,
                delimiter=delimiter,
                quotechar='"',
                skipinitialspace=True,
                on_bad_lines='skip',
                chunksize=self.CSV_CHUNK_SIZE,
            )
            for chunk in reader:
                chunks.append(chunk)
        except Exception:
            return None

        if not chunks:
            return None
        df = pd.concat(chunks, ignore_index=True)
        return self.clean_column_names(df)
    
    def _merge_two_columns_into_one(
        self, df: pd.DataFrame, col_a: str, col_b: str, keep_name: str
    ) -> pd.DataFrame:
        """Merge two columns (e.g. Brand and Brands) into one; keep keep_name, coalesce values, drop the other."""
        other = col_b if keep_name == col_a else col_a
        if keep_name not in df.columns or other not in df.columns:
            return df
        out = df.copy()
        out[keep_name] = out[keep_name].fillna(out[other])
        out = out.drop(columns=[other])
        return out

    def merge_dataframes(self, dataframes: List[pd.DataFrame]) -> pd.DataFrame:
        """
        Merge multiple DataFrames, handling column mismatches.
        
        Args:
            dataframes: List of DataFrames to merge
            
        Returns:
            Merged DataFrame
        """
        if not dataframes:
            return pd.DataFrame()
            
        if len(dataframes) == 1:
            return dataframes[0]
            
        # Use pandas.concat with sort=False to preserve column order
        # This automatically handles missing columns by filling with NaN
        merged_df = pd.concat(dataframes, ignore_index=True, sort=False)

        return merged_df

    # Allowed aggregation functions for pivot/aggregate step (must be valid pandas groupby.agg() names)
    AGG_FUNCS = frozenset({'sum', 'mean', 'count', 'first', 'last', 'min', 'max', 'median', 'std'})

    def _aggregate_by_group(
        self,
        df: pd.DataFrame,
        group_by: List[str],
        aggregate_columns: Optional[Dict[str, str]] = None,
        aggregate_default: Optional[str] = None,
    ) -> pd.DataFrame:
        """
        Aggregate a DataFrame by grouping on group_by columns.

        Pivot spec:
          - group_by: columns that define the groups (rows).
          - aggregate_columns: dict of column_name -> agg_func for value columns (sum, mean, count, first, etc.).
          - aggregate_default: agg for columns not in aggregate_columns. If None: sum for numeric, first for others.
        """
        missing = [c for c in group_by if c not in df.columns]
        if missing:
            self.errors.append(f"group_by columns not in data: {missing}")
            return df
        value_cols = [c for c in df.columns if c not in group_by]
        if not value_cols:
            return df.drop_duplicates(subset=group_by).reset_index(drop=True)
        agg_spec = aggregate_columns if aggregate_columns is not None else self.aggregate_columns
        default_agg = aggregate_default if aggregate_default is not None else self.aggregate_default
        agg_dict = {}
        for c in value_cols:
            if agg_spec and c in agg_spec:
                func = agg_spec[c]
                if func in self.AGG_FUNCS:
                    agg_dict[c] = func
                else:
                    agg_dict[c] = default_agg or ('sum' if pd.api.types.is_numeric_dtype(df[c]) else 'first')
            elif default_agg:
                agg_dict[c] = default_agg
            else:
                agg_dict[c] = 'sum' if pd.api.types.is_numeric_dtype(df[c]) else 'first'
        return df.groupby(group_by, as_index=False).agg(agg_dict)

    def merge_subdirectory(self, subdir_path: Path) -> bool:
        """
        Merge all supported files (Excel and CSV) in a subdirectory.

        Uses per-subdirectory config (columns, group_by, max_rows_before_aggregate) when available.
        """
        subdir_config = self.subdirectory_config.get(subdir_path.name, {})
        columns_for_this = subdir_config.get('columns') or self.columns_to_export
        group_by_for_this = subdir_config.get('group_by') if 'group_by' in subdir_config else self.group_by
        max_rows_for_this = subdir_config.get('max_rows_before_aggregate') if 'max_rows_before_aggregate' in subdir_config else self.max_rows_before_aggregate
        aggregate_columns_for_this = subdir_config.get('aggregate_columns') if 'aggregate_columns' in subdir_config else self.aggregate_columns
        aggregate_default_for_this = subdir_config.get('aggregate_default') if 'aggregate_default' in subdir_config else self.aggregate_default

        supported_files = self.get_supported_files(subdir_path)
        
        if not supported_files:
            print(f"  No supported files found in {subdir_path.name}")
            return True
            
        print(f"  Found {len(supported_files)} supported files")
        
        # Check if there's only one file - no merge needed
        if len(supported_files) == 1:
            print(f"  ℹ️  Only one file found in {subdir_path.name} - no merge needed")
            self.skipped_single_file += 1
            return True
        
        # Read all supported files (column cleaning is applied during read)
        dataframes = []
        file_paths_read = []
        for file_path in supported_files:
            df, method_used = self.read_file_with_method(file_path)
            if df is not None:
                dataframes.append(df)
                file_paths_read.append(file_path)
                print(f"    ✓ Read {file_path.name} ({len(df)} rows, {len(df.columns)} columns) using {method_used}")
            else:
                print(f"    ✗ Failed to read {file_path.name}")

        if not dataframes:
            print(f"  No valid files to merge in {subdir_path.name}")
            return False

        # Use reference per subfolder: the file with the most columns in this subfolder
        if len(dataframes) > 1:
            max_col_idx = max(range(len(dataframes)), key=lambda i: len(dataframes[i].columns))
            if max_col_idx != 0:
                file_paths_read[0], file_paths_read[max_col_idx] = file_paths_read[max_col_idx], file_paths_read[0]
                dataframes[0], dataframes[max_col_idx] = dataframes[max_col_idx], dataframes[0]
                print(f"  ℹ️  Reference for this subfolder: {file_paths_read[0].name} ({len(dataframes[0].columns)} columns)")

        # Resolve same-file plural/singular conflicts (e.g. file has both "Brand" and "Brands")
        if self.validate_structure and len(dataframes) > 1:
            while True:
                conflict_file_idx = None
                conflict_pair = None
                for i, df in enumerate(dataframes):
                    pair = file_has_plural_singular_conflict(df.columns.tolist())
                    if pair is not None:
                        conflict_file_idx = i
                        conflict_pair = pair
                        break
                if conflict_file_idx is None:
                    break
                path = file_paths_read[conflict_file_idx]
                col_a, col_b = conflict_pair
                print(f"  ⚠️  File has both '{col_a}' and '{col_b}': {path.name}")
                if not sys.stdin.isatty():
                    self.errors.append(
                        f"Same-file column conflict in {path.name} (both '{col_a}' and '{col_b}'). "
                        "Run interactively to choose which to keep, or fix the file manually."
                    )
                    return False
                while True:
                    try:
                        choice = input(
                            f"    Which name should we keep? (1) {col_a}  (2) {col_b}  (3) Abort [1/2/3]: "
                        ).strip() or "3"
                        if choice == "3":
                            self.errors.append(
                                f"User aborted: same-file column conflict in {path.name}"
                            )
                            return False
                        if choice == "1":
                            keep_name = col_a
                            break
                        if choice == "2":
                            keep_name = col_b
                            break
                    except EOFError:
                        self.errors.append(
                            f"Aborted (no input): same-file column conflict in {path.name}"
                        )
                        return False
                dataframes[conflict_file_idx] = self._merge_two_columns_into_one(
                    dataframes[conflict_file_idx], col_a, col_b, keep_name
                )
                print(f"    ✓ Merged columns into '{keep_name}'")

            # Apply cross-file plural/singular correction (majority wins)
            renames = compute_canonical_plural_singular_renames(
                [df.columns.tolist() for df in dataframes]
            )
            if renames:
                for i, df in enumerate(dataframes):
                    apply = {k: v for k, v in renames.items() if k in df.columns}
                    if apply:
                        dataframes[i] = df.rename(columns=apply)
                        print(f"    ✓ Normalized column names (plural/singular): {list(apply.keys())} → canonical")

            # Validate column structure after corrections (allow_missing_columns: merge leaves them blank)
            result = validate_column_structure(
                file_paths_read, dataframes, allow_missing_columns=True
            )
            if not result.success:
                self.errors.append(
                    f"Structure validation failed for subdirectory {subdir_path.name}"
                )
                print(f"  ✗ Column structure mismatch (merge skipped)")
                print(format_validation_report(result))
                return False

        # Warn when any file has columns not in the reference (extra columns in merged file)
        if len(dataframes) > 1:
            reference_set = set(dataframes[0].columns)
            extra_to_files: Dict[str, List[str]] = {}
            for i, df in enumerate(dataframes):
                for col in df.columns:
                    if col not in reference_set:
                        extra_to_files.setdefault(col, []).append(file_paths_read[i].name)
            if extra_to_files:
                print(f"  ⚠️  Extra columns (not in reference) – please verify they are not typos:")
                for col in sorted(extra_to_files.keys()):
                    files_str = ", ".join(extra_to_files[col])
                    print(f"      Column \"{col}\": in files: [{files_str}]")

        # Restrict to columns when provided (global or per-subdir)
        if columns_for_this is not None:
            for i in range(len(dataframes)):
                dataframes[i] = dataframes[i].reindex(columns=columns_for_this)
            in_any = set()
            for df in dataframes:
                for c in columns_for_this:
                    if c in df.columns and df[c].notna().any():
                        in_any.add(c)
            missing = set(columns_for_this) - in_any
            if missing:
                print(f"  ⚠️  Requested columns not present in any file: {sorted(missing)}")

        # Merge the dataframes
        merged_df = self.merge_dataframes(dataframes)

        # Optionally aggregate by group_by when row count exceeds threshold (pivot: group + value columns + agg)
        if group_by_for_this and max_rows_for_this is not None and len(merged_df) > max_rows_for_this:
            before = len(merged_df)
            merged_df = self._aggregate_by_group(
                merged_df,
                group_by_for_this,
                aggregate_columns=aggregate_columns_for_this,
                aggregate_default=aggregate_default_for_this,
            )
            agg_msg = f" (agg: {aggregate_columns_for_this})" if aggregate_columns_for_this else ""
            print(f"  📊 Aggregated by {group_by_for_this}{agg_msg}: {before} rows → {len(merged_df)} rows")
            # Pivot output: keep only group_by and aggregate_columns; drop any other columns
            pivot_cols = list(group_by_for_this) + (list(aggregate_columns_for_this.keys()) if aggregate_columns_for_this else [])
            pivot_cols = [c for c in pivot_cols if c in merged_df.columns]
            if pivot_cols:
                merged_df = merged_df[pivot_cols]

        # Sort before save if sort_by is configured
        sort_by_for_this = subdir_config.get('sort_by') if 'sort_by' in subdir_config else self.sort_by
        sort_ascending_for_this = subdir_config.get('sort_ascending') if 'sort_ascending' in subdir_config else self.sort_ascending
        if sort_by_for_this:
            sort_cols = [c for c in (sort_by_for_this if isinstance(sort_by_for_this, list) else [sort_by_for_this]) if c in merged_df.columns]
            if sort_cols:
                merged_df = merged_df.sort_values(by=sort_cols, ascending=sort_ascending_for_this).reset_index(drop=True)

        # Enforce output column order from config (columns list) so the saved file matches the order you specify
        if columns_for_this:
            order_from_config = [c for c in columns_for_this if c in merged_df.columns]
            extra = [c for c in merged_df.columns if c not in columns_for_this]
            if order_from_config or extra:
                merged_df = merged_df[order_from_config + extra]

        # Create output filename
        output_filename = f"{subdir_path.name}_merged.xlsx"
        output_path = subdir_path / output_filename
        
        # Save the merged file
        try:
            merged_df.to_excel(output_path, index=False)
            print(f"  ✓ Created merged file: {output_filename} ({len(merged_df)} rows, {len(merged_df.columns)} columns)")
            self.total_files_merged += len(dataframes)
            return True
        except Exception as e:
            self.errors.append(f"Error saving merged file {output_path}: {str(e)}")
            print(f"  ✗ Failed to save merged file: {str(e)}")
            return False
    
    def _get_subdirectories_to_process(self) -> List[Path]:
        """Return list of subdirectories to process, respecting subdir_include and subdir_exclude."""
        all_subdirs = sorted([d for d in self.root_directory.iterdir() if d.is_dir()])
        if self.subdir_include is not None:
            include_set = set(self.subdir_include)
            all_subdirs = [d for d in all_subdirs if d.name in include_set]
        exclude_set = set(self.subdir_exclude)
        return [d for d in all_subdirs if d.name not in exclude_set]

    def _get_all_file_columns_for_subdir(self, subdir_path: Path) -> List[Tuple[Path, List[str]]]:
        """
        Get (file_path, columns_list) for every file in this subdirectory that can be read.
        Uses the same read/cleaning as merge so names match. Returns [(path, columns), ...] or [].
        """
        supported_files = self.get_supported_files(subdir_path)
        if not supported_files:
            return []
        result = []
        for file_path in supported_files:
            df, _ = self.read_file_with_method(file_path)
            if df is not None:
                result.append((file_path, df.columns.tolist()))
        return result

    def _get_required_columns_for_subdir(self, subdir_path: Path) -> List[str]:
        """Return the union of columns required by config for this subdir (columns, group_by, aggregate_columns keys, sort_by)."""
        subdir_config = self.subdirectory_config.get(subdir_path.name, {})
        columns_for_this = subdir_config.get('columns') or self.columns_to_export or []
        group_by_for_this = subdir_config.get('group_by') if 'group_by' in subdir_config else self.group_by or []
        aggregate_columns_for_this = subdir_config.get('aggregate_columns') if 'aggregate_columns' in subdir_config else self.aggregate_columns or {}
        sort_by_for_this = subdir_config.get('sort_by') if 'sort_by' in subdir_config else self.sort_by or []
        if not isinstance(columns_for_this, list):
            columns_for_this = [columns_for_this]
        if not isinstance(group_by_for_this, list):
            group_by_for_this = [group_by_for_this] if group_by_for_this else []
        if not isinstance(sort_by_for_this, list):
            sort_by_for_this = [sort_by_for_this] if sort_by_for_this else []
        required = set(columns_for_this) | set(group_by_for_this) | set(aggregate_columns_for_this.keys()) | set(sort_by_for_this)
        return list(required)

    def _validate_required_columns(self, subdirs: List[Path]) -> Tuple[bool, List[str], List[Path]]:
        """
        Pre-flight check: for each subdir, ensure all required columns exist in every file to be merged.
        If any file is missing any required column, report it and abort.
        Returns (True, [], []) if all OK, else (False, list of error messages, list of failing subdir paths).
        """
        errors = []
        failing_subdirs: List[Path] = []
        for subdir_path in subdirs:
            required = self._get_required_columns_for_subdir(subdir_path)
            if not required:
                continue
            file_columns_list = self._get_all_file_columns_for_subdir(subdir_path)
            if not file_columns_list:
                errors.append(
                    f"Subdirectory '{subdir_path.name}': no supported files or could not read any file. "
                    "Cannot verify required columns."
                )
                failing_subdirs.append(subdir_path)
                continue
            required_set = set(required)
            subdir_has_error = False
            for file_path, columns in file_columns_list:
                file_set = set(columns)
                missing = [c for c in required if c not in file_set]
                if missing:
                    subdir_has_error = True
                    errors.append(
                        f"Subdirectory '{subdir_path.name}', file '{file_path.name}': "
                        f"missing required columns: {sorted(missing)}"
                    )
            if subdir_has_error:
                failing_subdirs.append(subdir_path)
        return (len(errors) == 0, errors, failing_subdirs)

    def process_all_subdirectories(self) -> Dict[str, Any]:
        """
        Process all subdirectories in the root directory (or only those in subdir_include, excluding subdir_exclude).

        Returns:
            Dictionary with processing results
        """
        if not self.validate_directory():
            return {
                'success': False,
                'errors': self.errors,
                'processed_subdirs': 0,
                'total_files_merged': 0
            }

        subdirs = self._get_subdirectories_to_process()
        total_subdirs = len([d for d in self.root_directory.iterdir() if d.is_dir()])

        print(f"Processing directory: {self.root_directory}")
        print(f"Found {total_subdirs} subdirectories" + (f", processing {len(subdirs)}" if len(subdirs) != total_subdirs else ""))
        print("-" * 50)

        # Pre-flight: verify required columns exist in reference file of each subdirectory
        if subdirs:
            ok, validation_errors, failing_subdirs = self._validate_required_columns(subdirs)
            if not ok:
                for msg in validation_errors:
                    self.errors.append(msg)
                    print(f"  ✗ {msg}")
                print()
                # Print columns per file for failing subdirs so user can build column_aliases in YAML
                if failing_subdirs:
                    print("Columns per file (for building column_aliases in YAML):")
                    print("-" * 50)
                    for subdir_path in failing_subdirs:
                        file_columns_list = self._get_all_file_columns_for_subdir(subdir_path)
                        if not file_columns_list:
                            continue
                        print(f"  Subdirectory: {subdir_path.name}")
                        for file_path, columns in file_columns_list:
                            print(f"    {file_path.name}: {columns}")
                        print()
                print("Aborting: fix missing columns or config and run again.")
                return {
                    'success': False,
                    'errors': self.errors,
                    'processed_subdirs': 0,
                    'total_files_merged': 0,
                    'total_subdirs': len(subdirs)
                }

        for subdir in subdirs:
            print(f"Processing subdirectory: {subdir.name}")
            success = self.merge_subdirectory(subdir)
            if success:
                self.processed_subdirs += 1
            print()
        
        return {
            'success': True,
            'processed_subdirs': self.processed_subdirs,
            'total_files_merged': self.total_files_merged,
            'skipped_single_file': self.skipped_single_file,
            'errors': self.errors,
            'total_subdirs': len(subdirs)
        }

    def print_summary(self, results: Dict[str, Any]):
        """Print a summary of the processing results."""
        print("=" * 50)
        print("PROCESSING SUMMARY")
        print("=" * 50)
        
        if results['success']:
            print(f"✓ Successfully processed {results['processed_subdirs']} out of {results['total_subdirs']} subdirectories")
            print(f"✓ Merged {results['total_files_merged']} files")
            if results.get('skipped_single_file', 0) > 0:
                print(f"ℹ️  Skipped {results['skipped_single_file']} subdirectories with only one file (no merge needed)")
        else:
            print("✗ Processing failed")
            
        if results['errors']:
            print(f"\n⚠️  {len(results['errors'])} errors encountered:")
            for error in results['errors']:
                print(f"  - {error}")
        
        print("=" * 50)
